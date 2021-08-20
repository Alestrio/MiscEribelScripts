#  Copyright (c) 2021-2022. Alexis LEBEL
#  All rights reserved
#  This code belongs exclusively to it's author, and cannot be reproduced legally without it's author's authorization.
#  This code is made available to use to ERIBEL.
#  If you find this file publicly available anywhere else than on the authors' Github, please warn it's author.

import json
import sys
import re
import os

from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.graph_client import GraphClient

import cloudsync
import msal

from openpyxl import load_workbook

conf_filename = "auto_rename_conf.json"
workbook = None
sheet = None


def get_conf_from_json():
  """Fetch config from json config file"""
  with open('auto_rename_conf.json') as file:
    conf =  json.loads(file.read())
  if conf:
    return conf
  else:
    raise Exception("Unable to fetch configuration")
	
def acquire_token_func():
	"""
	Acquire token via MSAL
	"""
	conf = get_conf_from_json()
	tenant_name = conf['tenant_name']
	authority_url = f'https://login.microsoftonline.com/{tenant_name}'
	app = msal.ConfidentialClientApplication(
		authority=authority_url,
		client_id=conf['client_id'],
		client_credential=conf['client_secret']
	)
	token = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
	return token

def copy_excel_from_sharepoint():
	"""Copy excel from sharepoint"""
	conf = get_conf_from_json()
	token = acquire_token_func
	client = GraphClient(token)
	tenant_prefix = 'eribel354'
	file_abs_url = conf['file_abs_url']
	file_item = client.shares.by_url(file_abs_url).drive_item.get().execute_query()
	with open('./temp_excel.xlsm', 'wb') as tempxl:
		file_item.download(tempxl).execute_query()

def load_spreadsheet():
  """Loads spreadsheet"""
  workbook = load_workbook(filename='./temp_excel.xlsm')


def set_user_defined_sheet_name():
  """Ask the user which sheet to use (Machine Nr - Invoice Nr) and set it as active"""
  probable_sheets = []

  machine_nr = input('Veuillez entrer un numéro de machine (Vide si recherche par C.) : ')
  invoice_nr = input('Veuillez entrer un numéro de C. : ')
  sheets = workbook.sheetnames

  # Research amongs all sheets (There can be a lot)
  for ii in sheets:
    if ii.contains(machine_nr) and ii.contains(invoice_nr):
      sheet = workbook['ii']
      break
    elif ii.contains(machine_nr):
      probable_sheets.append(ii)
    elif ii.contains(invoice_nr):
      probable_sheets.append(ii)

  # If no exact corresponding sheet is found
  if not sheet and probable_sheets != []:
    print('Aucune feuille ne correspond totalement à votre recherche, mais certaines s\'en rapprochent :')
    i = 0
    for ii in probable_sheets:
      print(f'{i} : {ii}')
      i+=1
    print('99 pour quitter')
    choice = input('Faites un choix :')

    # Let the user exit the script
    if choice == 'q':
      sys.exit()
    else:
      sheet = workbook[probable_sheets[int(choice)]]

def get_old_numbers_column():
  #Get old numbers column
  cells_to_check = sheet['A17', 'G24']
  for ii in cells_to_check:
    if ii.value == 'ancienne':
      return ii.column_letter
  raise Exception('No column named "ancienne" in cell range')

def get_new_numbers_columns():
  #Get new numbers columns
  new_numbers_columns = []
  cells_to_check = sheet['A17', 'BK24']
  for ii in cells_to_check:
    if ii.value == 'nouvelle':
      new_numbers_columns.append(ii.column_letter)
  if new_numbers_columns != []:
    return new_numbers_columns
  else:
    raise Exception('No column named "nouvelle" in cell range')

def parse_columns():
  """Creates array of tuples for numbers replacement"""
  # Fetch columns letters
  old_column_letter = get_old_numbers_column()
  new_columns_letters = get_new_numbers_columns()

  # Fetch old numbers
  old_column_cells = [f'{old_column_letter}15', f'{old_column_letter}120']
  old_numbers_cells = []
  for ii in old_column_cells:
    if re.search('[0-9]{7}'):
      old_numbers_cells.append(ii)

  # Fetch all new numbers (for each format)
  # Creates an array of arrays of cells ([[CellA, CellB], [CellC, CellD]])
  for ii in old_numbers_cells:
    new_numbers_cells_array = []
    for ij in new_columns_letters:
      new_columns_cells = [f'{ij}15', f'{ij}120']
      new_numbers_cells = []
      for jj in new_columns_cells:
        if re.search('[0-9]{7}'):
          new_numbers_cells.append(ii)
    new_numbers_cells_array.append(new_numbers_cells)

  # Combines all the cells :
  # Creates an array of tuples, easier to work with..
  work_tuples = []
  for ii in old_numbers_cells:
    # First we create an array..
    work_array = [int(ii.value)]
    # Then we fill it..
    for ij in new_numbers_cells_array:
      # Using the current position in the old cells array...
      work_array.append(ij[old_numbers_cells.index(ii)])
    # Finally, the array is parsed as a tuple and added to the list
    work_tuples.append(tuple(work_array))

    return work_tuples

def get_directory_definition():
  """Recursively replaces old numbers with new numbers"""
  work_tuples = parse_columns()

  # Asking directory to work into :
  dir = input('Veuillez entrer le chemin de travail (Format absolu : "C:/...)" : ')
  # Listing equipement directories :
  equipt_dirs = []
  for ii in os.listdir(dir):
    if ii.contains('EQUIPT'):
      equipt_dirs.append(dir + '\\' + ii)

  # Testing data coherence :
  if not len(work_tuples[1])-1 == len(equipt_dirs):
    raise Exception("Data coherence test failed, not enough EQUIPTS for numbers")

  # Associating a column to a folder :
  # Printing all equipt folders :
  print("""La prochaine étape consiste à associer un équipement à une colonne de numéros. \n
  Veuillez selectionner l'ordre des équipements en rentrant les numéros dans le bon ordre (ex : 132)""")
  print('-------------------------')
  for ii in equipt_dirs:
    print(f'| {equipt_dirs.index(ii)} | {ii} |')
  print('-------------------------')
  # Input and parsing :
  while True:
    order = input('Ordre : ')
    if len(order) == len(equipt_dirs):
      break

  # Ordering equipement according to user input
  ordered_equipts = []
  for ii in order:
    ordered_equipts.append(equipt_dirs[int(ii)])

  return ordered_equipts

def rename_file(file_path, equipt_nr):
  """Renames a file with it's corresponding new name"""
  work_tuples = parse_columns()
  # Regex used to get differents parts of the file path
  path_regex = re.compile(r'(?P<path>[\w\\:]*)\\(?P<filename>[\w]*).(?P<extension>[\w].)')
  # Match object containing the different parts of the file path
  match = path_regex.search(file_path)

  # Getting the right file to rename
  associated_nr = 0
  for ii in work_tuples:
    if match.group('filename') == ii[0]:
      associated_nr = ii[equipt_nr+1]

  # Renaming the file
  os.rename(file_path, match.group('path')+'\\'+associated_nr+'.'+match.group('extension'))

def iterate_dir(dir_path:str, files, equipt_nr):
  """Recursive function iterating a directory"""
  for ii in os.listdir(dir_path):
    if os.path.isdir(ii):
      iterate_dir(ii)
    elif re.search('[0-9]{7}', ii):
      rename_file(ii, equipt_nr)
    else:
      print('not editing : ' + ii)


def recursively_rename_files():
  """Recursively replaces old numbers with new numbers"""
  ordered_equipts = get_directory_definition()

  # Iterates each equipement folder
  for ii in ordered_equipts:
    iterate_dir(ii, ordered_equipts.index(ii))

def open_file_link_manager(file):
  """Opens file-link manager for each renamed file"""
  pass

def create_numbers_table():
  """Creates a table which contains all the numbers"""
  work_tuples = parse_columns()
  print('\n\n\n ----- Tableau récapitulatif -----')
  print('-----------------------')
  for ii in work_tuples:
    line = '|'
    for ij in ii:
      line += ' ij |'
    print(line)
  print('-----------------------')

def delete_temp_excel():
  """Deletes temporary excel"""
  os.remove('./temp_excel.xlsx')

if __name__ == '__main__':
  #Main function
  copy_excel_from_sharepoint()
  load_spreadsheet()
  set_user_defined_sheet_name()
  recursively_rename_files()
  create_numbers_table()
  delete_temp_excel()
  